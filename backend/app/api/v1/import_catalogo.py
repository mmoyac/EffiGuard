"""
Importación Excel del catálogo producto → variante → unidad.

Dos principios que gobiernan el archivo completo:

1. **El Excel mantiene el catálogo; el inventario sólo se mueve por movimientos.**
   Un `stock_actual` declarado nunca se escribe directo: se traduce en un log de
   apertura al crear y en un `ajuste` al actualizar. El template está pensado para
   uso recurrente, y ahí una escritura muda sería un bug silencioso sobre el dato
   que el sistema existe para custodiar.

2. **Celda vacía nunca borra.** En una fila que actualiza, lo que no viene se
   conserva. La mitad de las reimportaciones son planillas parciales con dos
   columnas; interpretarlas como orden de borrado sería destruir datos por omisión.
"""
import io
from decimal import Decimal, InvalidOperation

import openpyxl
import openpyxl.styles
import openpyxl.utils
from fastapi import APIRouter, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select

from app.core.dependencies import CurrentToken, DBSession
from app.models.asset_family import AssetFamily
from app.models.brand import Brand
from app.models.codigo import TIPOS_DE_UNIDAD, TIPOS_VALIDOS, Codigo
from app.models.producto import Producto
from app.models.variante import Variante
from app.repositories.inventory_log import InventoryLogRepository
from app.repositories.producto import ProductoRepository
from app.repositories.proveedor import ProveedorRepository
from app.repositories.ubicacion import UbicacionRepository
from app.repositories.unidad import UnidadRepository
from app.repositories.variante import VarianteRepository
from app.schemas.common import UNIDADES_VALIDAS
from app.schemas.catalogo import CodigoCreate, VarianteCreate
from app.services import catalogo as svc

router = APIRouter(prefix="/catalogo/import", tags=["Catálogo Import"])

# Se leen por índice: las columnas nuevas SIEMPRE se agregan al final, porque
# insertarlas en medio rompería en silencio los archivos ya descargados.
_COLUMNS = [
    "producto",
    "variante",
    "familia",
    "marca",
    "unidad",
    "stock_actual",
    "stock_minimo",
    "precio_compra",
    "valor_reposicion",
    "dias_max_prestamo",
    "codigos",
    "cantidad_unidades",
    "ubicacion",
]

def _ejemplos(consumible: str, prestable: str) -> list[list[str]]:
    """Filas de ejemplo con las familias REALES del tenant.

    Hardcodear "Consumible"/"Herramienta" hacía que el template se descargara ya
    roto para cualquier tenant que hubiera nombrado distinto sus familias: el
    usuario lo subía tal cual y cada fila se rechazaba por familia inexistente.
    """
    return [
        # Un consumible con dos proveedores y dos empaques de distinto contenido:
        # el caso que motivó todo el rediseño.
        ["Tornillo autoperforante", "6x40 zincado", consumible, "", "unidad", "500", "200", "12", "", "",
         "7801234567890:proveedor:1:Sodimac;17801234567890:empaque:100:Sodimac;"
         "7809876543210:proveedor:1:Construmart;17809876543210:empaque:250:Construmart", "", "RACK-A/N2/P3"],
        # Otra variante del mismo producto: repetir el nombre agrupa, no duplica.
        ["Tornillo autoperforante", "8x60 zincado", consumible, "", "unidad", "300", "150", "18", "", "",
         "7801234567906:proveedor:1:Sodimac", "", "RACK-A/N2/P4"],
        # Herramienta sin etiquetar: cantidad_unidades crea los ejemplares.
        ["Taladro percutor GSB-13RE", "", prestable, "Bosch", "unidad", "", "1", "", "180000", "7",
         "4053423205718:fabricante", "3", "RACK-C/N1/P1"],
        # Herramienta ya etiquetada: una fila por ejemplar, mismo producto.
        ["Esmeril angular GWS-850", "", prestable, "Bosch", "unidad", "", "1", "", "145000", "7",
         "4059952533445:fabricante;QR-00417:propio", "", "RACK-C/N1/P2"],
        ["Esmeril angular GWS-850", "", prestable, "Bosch", "unidad", "", "1", "", "145000", "7",
         "QR-00418:propio;SN-8842190337:serie_fabrica", "", "RACK-C/N1/P2"],
    ]


@router.get("/template")
async def download_template(token: CurrentToken, session: DBSession):
    """
    Template del catálogo: una fila por variante.

    `stock_actual` y `cantidad_unidades` se exportan **vacías** a propósito. Si el
    dato no está, no hay forma de pisar inventario ni duplicar ejemplares al
    reimportar. Quien quiera usarlas para una carga inicial las escribe a mano.
    """
    var_repo = VarianteRepository(session, token.tenant_id)
    filas = await var_repo.listar(limit=5000)

    ubic_repo = UbicacionRepository(session, token.tenant_id)
    ubicaciones = {u.id: u for u in await ubic_repo.list_all()}

    brands_result = await session.execute(
        select(Brand).where(Brand.tenant_id == token.tenant_id)
    )
    brands = {b.id: b.nombre for b in brands_result.scalars().all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo"

    header_fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = openpyxl.styles.Font(bold=True, color="E5E7EB")
    for col, header in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    if filas:
        for row_idx, (variante, _total, _disp) in enumerate(filas, start=2):
            producto = variante.producto
            ubic = ubicaciones.get(variante.ubicacion_id) if variante.ubicacion_id else None
            codigos = ";".join(
                _formatear_codigo(c) for c in variante.codigos
            )
            ws.cell(row=row_idx, column=1, value=producto.nombre)
            ws.cell(row=row_idx, column=2, value=variante.nombre)
            ws.cell(row=row_idx, column=3, value=producto.family.nombre)
            ws.cell(row=row_idx, column=4, value=brands.get(producto.brand_id, "") if producto.brand_id else "")
            ws.cell(row=row_idx, column=5, value=variante.unidad)
            ws.cell(row=row_idx, column=6, value="")   # stock_actual — deliberadamente vacía
            ws.cell(row=row_idx, column=7, value=float(variante.stock_minimo or 0))
            ws.cell(row=row_idx, column=8, value=float(variante.precio_compra) if variante.precio_compra else "")
            ws.cell(row=row_idx, column=9, value=float(variante.valor_reposicion) if variante.valor_reposicion else "")
            ws.cell(row=row_idx, column=10, value=variante.dias_max_prestamo or "")
            ws.cell(row=row_idx, column=11, value=codigos)
            ws.cell(row=row_idx, column=12, value="")  # cantidad_unidades — deliberadamente vacía
            ws.cell(row=row_idx, column=13, value=f"{ubic.rack}/{ubic.nivel}/{ubic.posicion}" if ubic else "")
    else:
        familias = (
            await session.execute(
                select(AssetFamily).where(AssetFamily.tenant_id == token.tenant_id)
            )
        ).scalars().all()
        consumible = next(
            (f.nombre for f in familias if f.comportamiento == "consumible"), "Consumible"
        )
        prestable = next(
            (f.nombre for f in familias if f.comportamiento == "prestable"), "Herramienta"
        )
        for row_idx, row_data in enumerate(_ejemplos(consumible, prestable), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    widths = [30, 22, 18, 14, 10, 13, 13, 14, 17, 18, 55, 18, 20]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=catalogo.xlsx"},
    )


@router.post("")
async def import_catalogo(
    token: CurrentToken,
    session: DBSession,
    file: UploadFile = File(...),
    dry_run: bool = Query(False, description="Si true, valida sin modificar datos"),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo debe ser .xlsx")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se pudo leer el archivo Excel")

    rows = list(wb.active.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo no tiene filas de datos")

    tenant_id = token.tenant_id
    prod_repo = ProductoRepository(session, tenant_id)
    var_repo = VarianteRepository(session, tenant_id)
    prov_repo = ProveedorRepository(session, tenant_id)
    ubic_repo = UbicacionRepository(session, tenant_id)
    unidad_repo = UnidadRepository(session, tenant_id)
    log_repo = InventoryLogRepository(session, tenant_id)

    familias_result = await session.execute(
        select(AssetFamily).where(AssetFamily.tenant_id == tenant_id)
    )
    familias = {f.nombre.strip().lower(): f for f in familias_result.scalars().all()}

    brands_result = await session.execute(select(Brand).where(Brand.tenant_id == tenant_id))
    marcas = {b.nombre.strip().lower(): b.id for b in brands_result.scalars().all()}

    # Códigos ya registrados, con la identidad de su dueño. Es lo que permite
    # distinguir "este código ya cuelga de esta misma variante" (idempotente, y por
    # eso reimportar no falla) de "este código es de otro item" (error de fila).
    codigos_db: dict[str, Codigo] = {}
    codigos_vistos: dict[str, tuple[str, object]] = {}

    codigos_result = await session.execute(
        select(Codigo, Producto.nombre, Variante.nombre)
        .outerjoin(Variante, Codigo.variante_id == Variante.id)
        .outerjoin(Producto, Variante.producto_id == Producto.id)
        .where(Codigo.tenant_id == tenant_id)
    )
    for codigo_obj, prod_nombre, var_nombre in codigos_result.all():
        codigos_db[codigo_obj.codigo] = codigo_obj
        if codigo_obj.variante_id and prod_nombre and var_nombre:
            codigos_vistos[codigo_obj.codigo] = (
                "variante",
                (prod_nombre.strip().lower(), var_nombre.strip().lower()),
            )
        else:
            codigos_vistos[codigo_obj.codigo] = ("unidad", codigo_obj.codigo)

    errores: list[dict] = []
    advertencias: list[dict] = []
    ajustes_stock: list[dict] = []
    creados = {
        "productos": 0, "variantes": 0, "unidades": 0,
        "codigos": 0, "proveedores": 0, "ubicaciones": 0,
    }

    # Caches de lo planificado en ESTE archivo, para que la fila 3 vea lo que
    # creó la fila 2 aun en dry_run, donde nada se persiste.
    productos_cache: dict[str, object] = {}
    variantes_cache: dict[tuple[str, str], object] = {}
    nombres_planificados: set[str] = set()
    variantes_planificadas: set[tuple[str, str]] = set()
    proveedores_planificados: set[str] = set()
    # Por variante y no por fila: las tres filas de un mismo esmeril tocan una
    # sola variante, y contarlas tres veces exageraría el impacto del archivo.
    variantes_tocadas: set[tuple[str, str]] = set()

    async def resolver_ubicacion(texto: str | None) -> int | None:
        """Acepta 'RACK/NIVEL/POSICION'. Se crea si no existe, igual que antes:
        una ubicación es una etiqueta de dónde está algo, no configuración."""
        if not texto:
            return None
        partes = [p.strip().upper() for p in str(texto).split("/") if p.strip()]
        if len(partes) != 3:
            raise ValueError("la ubicación requiere rack/nivel/posición")
        existente = await ubic_repo.get_by_posicion(*partes)
        if existente:
            return existente.id
        creados["ubicaciones"] += 1
        if dry_run:
            return None
        nueva = await ubic_repo.create(rack=partes[0], nivel=partes[1], posicion=partes[2])
        return nueva.id

    async def resolver_proveedor(nombre: str | None) -> int | None:
        if not nombre:
            return None
        existente = await prov_repo.get_by_nombre(nombre)
        if existente:
            return existente.id
        # En dry_run nada se persiste, así que sin este registro de lo planificado
        # el mismo proveedor se contaría una vez por fila que lo menciona.
        clave = nombre.strip().lower()
        if clave in proveedores_planificados:
            return None
        proveedores_planificados.add(clave)
        creados["proveedores"] += 1
        if dry_run:
            return None
        nuevo = await prov_repo.create(nombre=nombre.strip())
        return nuevo.id

    for fila_num, row in enumerate(rows[1:], start=2):
        def col(idx: int) -> str | None:
            if idx >= len(row) or row[idx] is None:
                return None
            s = str(row[idx]).strip()
            return s or None

        producto_nombre = col(0)
        variante_nombre = col(1)
        familia_nombre = col(2)
        marca_nombre = col(3)
        unidad_raw = col(4)
        stock_actual_raw = col(5)
        stock_minimo_raw = col(6)
        precio_raw = col(7)
        valor_rep_raw = col(8)
        dias_raw = col(9)
        codigos_raw = col(10)
        cantidad_unidades_raw = col(11)
        ubicacion_raw = col(12)

        if not any([producto_nombre, variante_nombre, familia_nombre]):
            continue

        if not producto_nombre:
            errores.append({"fila": fila_num, "motivo": "producto es obligatorio"})
            continue
        if not familia_nombre:
            errores.append({"fila": fila_num, "motivo": "familia es obligatoria"})
            continue

        familia = familias.get(familia_nombre.strip().lower())
        if familia is None:
            errores.append({"fila": fila_num, "motivo": f"familia '{familia_nombre}' no existe en el sistema"})
            continue
        comportamiento = familia.comportamiento
        es_prestable = comportamiento == "prestable"

        # Variante vacía → homónima del producto, igual que el alta por formulario.
        variante_nombre = variante_nombre or producto_nombre
        clave_prod = producto_nombre.strip().lower()
        clave_var = (clave_prod, variante_nombre.strip().lower())

        # ── Parseo numérico ──────────────────────────────────────────────────
        unidad = (unidad_raw or "unidad").strip().lower()
        if unidad not in UNIDADES_VALIDAS:
            errores.append({"fila": fila_num, "motivo": f"unidad '{unidad_raw}' inválida. Opciones: {', '.join(UNIDADES_VALIDAS)}"})
            continue

        stock_actual = _num(stock_actual_raw)
        if stock_actual is False:
            errores.append({"fila": fila_num, "motivo": f"stock_actual '{stock_actual_raw}' no es un número válido"})
            continue
        stock_minimo = _num(stock_minimo_raw)
        if stock_minimo is False:
            errores.append({"fila": fila_num, "motivo": f"stock_minimo '{stock_minimo_raw}' no es un número válido"})
            continue
        precio_compra = _num(precio_raw)
        if precio_compra is False:
            errores.append({"fila": fila_num, "motivo": f"precio_compra '{precio_raw}' no es un número válido"})
            continue
        valor_reposicion = _num(valor_rep_raw)
        if valor_reposicion is False:
            errores.append({"fila": fila_num, "motivo": f"valor_reposicion '{valor_rep_raw}' no es un número válido"})
            continue
        dias_max = _entero(dias_raw)
        if dias_max is False:
            errores.append({"fila": fila_num, "motivo": f"dias_max_prestamo '{dias_raw}' no es un entero válido"})
            continue
        cantidad_unidades = _entero(cantidad_unidades_raw)
        if cantidad_unidades is False:
            errores.append({"fila": fila_num, "motivo": f"cantidad_unidades '{cantidad_unidades_raw}' no es un entero válido"})
            continue

        # ── Códigos: el tipo rutea el nivel ──────────────────────────────────
        try:
            codigos = _parsear_codigos(codigos_raw)
        except ValueError as e:
            errores.append({"fila": fila_num, "motivo": str(e)})
            continue

        de_variante = [c for c in codigos if c["tipo"] not in TIPOS_DE_UNIDAD and not (es_prestable and c["tipo"] == "propio")]
        de_unidad = [c for c in codigos if c not in de_variante]

        if cantidad_unidades and de_unidad:
            errores.append({"fila": fila_num, "motivo": "Declare cantidad_unidades o los códigos del ejemplar, no ambos"})
            continue
        if cantidad_unidades and not es_prestable:
            errores.append({"fila": fila_num, "motivo": "cantidad_unidades sólo aplica a familias prestables"})
            continue
        if de_unidad and not es_prestable:
            errores.append({"fila": fila_num, "motivo": "los códigos de ejemplar sólo aplican a familias prestables"})
            continue

        try:
            ubicacion_id = await resolver_ubicacion(ubicacion_raw)
        except ValueError as e:
            errores.append({"fila": fila_num, "motivo": str(e)})
            continue

        for c in codigos:
            c["proveedor_id"] = await resolver_proveedor(c.pop("proveedor", None))

        # ── Resolver producto ────────────────────────────────────────────────
        producto = productos_cache.get(clave_prod)
        if producto is None:
            producto = await prod_repo.get_por_nombre(producto_nombre)
        producto_es_nuevo = producto is None and clave_prod not in nombres_planificados

        if producto_es_nuevo:
            creados["productos"] += 1
            nombres_planificados.add(clave_prod)
            if not dry_run:
                nuevo = await prod_repo.create(
                    nombre=producto_nombre.strip(),
                    family_id=familia.id,
                    brand_id=marcas.get(marca_nombre.strip().lower()) if marca_nombre else None,
                )
                producto = await prod_repo.get_con_relaciones(nuevo.id)
                productos_cache[clave_prod] = producto

        # ── Resolver variante ────────────────────────────────────────────────
        variante = variantes_cache.get(clave_var)
        if variante is None and producto is not None:
            variante = await var_repo.get_por_nombre(producto.id, variante_nombre)
        variante_es_nueva = variante is None and clave_var not in variantes_planificadas

        if variante_es_nueva:
            creados["variantes"] += 1
            variantes_planificadas.add(clave_var)
            if not dry_run and producto is not None:
                data = VarianteCreate(
                    nombre=variante_nombre.strip(),
                    unidad=unidad,
                    # El stock NO se escribe acá: entra como movimiento más abajo.
                    stock_actual=Decimal(0),
                    stock_minimo=stock_minimo or Decimal(0),
                    precio_compra=precio_compra,
                    valor_reposicion=valor_reposicion,
                    dias_max_prestamo=dias_max,
                    ubicacion_id=ubicacion_id,
                )
                variante = await svc.crear_variante(producto, data, session, tenant_id)
                variantes_cache[clave_var] = variante
        elif variante is not None:
            # Actualización: celda vacía nunca borra.
            cambios = {}
            if unidad_raw is not None:
                cambios["unidad"] = unidad
            if stock_minimo is not None:
                cambios["stock_minimo"] = stock_minimo
            if precio_compra is not None:
                cambios["precio_compra"] = precio_compra
            if valor_reposicion is not None:
                cambios["valor_reposicion"] = valor_reposicion
            if dias_max is not None:
                cambios["dias_max_prestamo"] = dias_max
            if ubicacion_id is not None:
                cambios["ubicacion_id"] = ubicacion_id
            variantes_tocadas.add(clave_var)
            if not dry_run and cambios:
                await var_repo.update(variante, **cambios)

        # ── Códigos de la variante ───────────────────────────────────────────
        # Repetir el EAN del modelo en las tres filas de un mismo taladro NO es un
        # error: apunta al item que ya lo tiene. Sólo choca contra OTRO dueño.
        conflicto = None
        for c in de_variante:
            dueno = codigos_vistos.get(c["codigo"])
            if dueno is None:
                continue
            if dueno != ("variante", clave_var):
                conflicto = c["codigo"]
                break
        if conflicto:
            errores.append({"fila": fila_num, "motivo": f"el código '{conflicto}' ya está registrado en otro item"})
            continue

        for c in de_variante:
            if c["codigo"] in codigos_vistos:
                continue  # idempotente: ya cuelga de esta misma variante
            codigos_vistos[c["codigo"]] = ("variante", clave_var)
            creados["codigos"] += 1
            if not dry_run and variante is not None:
                await svc.agregar_codigo(
                    CodigoCreate(**c), session, tenant_id, variante_id=variante.id
                )

        # ── Unidades ─────────────────────────────────────────────────────────
        if cantidad_unidades:
            if not variante_es_nueva:
                advertencias.append({
                    "fila": fila_num,
                    "motivo": "cantidad_unidades se ignora: la variante ya existía, no se duplican ejemplares",
                })
            else:
                creados["unidades"] += cantidad_unidades
                creados["codigos"] += cantidad_unidades
                if not dry_run and variante is not None:
                    await svc.crear_unidades(
                        variante, cantidad_unidades, session, tenant_id,
                        comportamiento=comportamiento, ubicacion_id=ubicacion_id,
                    )

        if de_unidad:
            # Una fila por ejemplar. Su identidad es el primer código de nivel
            # unidad: si ya existe, la fila actualiza ese ejemplar en vez de crear
            # otro — es lo que hace que reimportar el mismo archivo no duplique.
            clave_unidad = de_unidad[0]["codigo"]
            dueno = codigos_vistos.get(clave_unidad)

            if dueno is not None and dueno[0] != "unidad":
                errores.append({"fila": fila_num, "motivo": f"el código '{clave_unidad}' ya está registrado en otro item"})
                continue

            unidad_existente = codigos_db.get(clave_unidad)
            if unidad_existente is not None and unidad_existente.unidad_id:
                # Ejemplar ya cargado: se actualiza su ubicación y nada más.
                if not dry_run and ubicacion_id is not None:
                    unidad_obj = await unidad_repo.get(unidad_existente.unidad_id)
                    if unidad_obj:
                        await unidad_repo.update(unidad_obj, ubicacion_id=ubicacion_id)
                advertencias.append({
                    "fila": fila_num,
                    "motivo": f"el ejemplar '{clave_unidad}' ya existía: se actualizó en vez de duplicarse",
                })
            else:
                creados["unidades"] += 1
                creados["codigos"] += len(de_unidad)
                codigos_vistos[clave_unidad] = ("unidad", clave_unidad)
                for c in de_unidad[1:]:
                    codigos_vistos[c["codigo"]] = ("unidad", clave_unidad)
                if not dry_run and variante is not None:
                    unidad_obj = await unidad_repo.create(
                        variante_id=variante.id, estado_id=1, ubicacion_id=ubicacion_id
                    )
                    for c in de_unidad:
                        await svc.agregar_codigo(
                            CodigoCreate(**c), session, tenant_id, unidad_id=unidad_obj.id
                        )

        # ── Stock: siempre como movimiento, nunca escritura muda ─────────────
        if stock_actual is not None and not es_prestable:
            vigente = variante.stock_actual if (variante is not None and not variante_es_nueva) else Decimal(0)
            if stock_actual != vigente:
                apertura = variante_es_nueva
                ajustes_stock.append({
                    "fila": fila_num,
                    "variante": f"{producto_nombre} · {variante_nombre}",
                    "de": float(vigente),
                    "a": float(stock_actual),
                    "tipo": "apertura" if apertura else "ajuste",
                })
                if not dry_run and variante is not None:
                    await var_repo.update(variante, stock_actual=stock_actual)
                    await log_repo.create(
                        variante_id=variante.id,
                        user_id=token.user_id,
                        tipo_movimiento="ajuste",
                        cantidad=abs(stock_actual - vigente),
                        observaciones=(
                            "Saldo de apertura: importación Excel"
                            if apertura
                            else f"Ajuste: {vigente} → {stock_actual}, importación Excel"
                        ),
                    )
        elif stock_actual is not None and es_prestable:
            advertencias.append({
                "fila": fila_num,
                "motivo": "stock_actual se ignora: el stock de una herramienta se deriva de sus unidades",
            })

    if not dry_run:
        await session.commit()

    return {
        "dry_run": dry_run,
        "productos_creados": creados["productos"],
        "variantes_creadas": creados["variantes"],
        "variantes_actualizadas": len(variantes_tocadas),
        "unidades_creadas": creados["unidades"],
        "codigos_creados": creados["codigos"],
        "proveedores_creados": creados["proveedores"],
        "ubicaciones_creadas": creados["ubicaciones"],
        "ajustes_stock": ajustes_stock,
        "advertencias": advertencias,
        "errores": errores,
    }


# ── helpers ──────────────────────────────────────────────────────────────────


def _formatear_codigo(c) -> str:
    """Serializa un código al formato de la columna, para el template descargable."""
    partes = [c.codigo, c.tipo, str(float(c.factor))]
    if c.proveedor:
        partes.append(c.proveedor.nombre)
    return ":".join(partes)


def _parsear_codigos(raw: str | None) -> list[dict]:
    """`codigo[:tipo[:factor[:proveedor]]]`, separados por `;`.

    El tipo determina a qué nivel cuelga el código, así que no hace falta una
    columna aparte para distinguir los de la variante de los del ejemplar.
    """
    if not raw:
        return []
    salida: list[dict] = []
    for entrada in str(raw).split(";"):
        entrada = entrada.strip()
        if not entrada:
            continue
        partes = [p.strip() for p in entrada.split(":")]
        codigo = partes[0].upper()
        if not codigo:
            continue
        tipo = (partes[1].lower() if len(partes) > 1 and partes[1] else "propio")
        if tipo not in TIPOS_VALIDOS:
            raise ValueError(f"tipo de código '{tipo}' inválido. Opciones: {', '.join(TIPOS_VALIDOS)}")
        factor = Decimal(1)
        if len(partes) > 2 and partes[2]:
            try:
                factor = Decimal(partes[2].replace(",", "."))
            except InvalidOperation:
                raise ValueError(f"factor '{partes[2]}' del código '{codigo}' no es un número válido")
            if factor <= 0:
                raise ValueError(f"el factor del código '{codigo}' debe ser mayor a 0")
        proveedor = partes[3] if len(partes) > 3 and partes[3] else None
        salida.append({
            "codigo": codigo,
            "tipo": tipo,
            "factor": factor,
            "proveedor": proveedor,
            # Deliberadamente vacío: el envase real puede ser caja, rollo, tambor o
            # saco, y el Excel no lo declara. Poner "caja" por defecto haría que un
            # rollo de 100 m se comprara como "3 cajas". Se asigna desde la interfaz.
            "nombre_empaque": None,
        })
    return salida


def _num(value: str | None):
    """None si la celda viene vacía (no borra), False si no es un número."""
    if value is None:
        return None
    try:
        return Decimal(str(value).strip().replace(",", ".")).quantize(Decimal("0.001"))
    except (InvalidOperation, ValueError, TypeError):
        return False


def _entero(value: str | None):
    if value is None:
        return None
    try:
        return int(float(str(value).strip().replace(",", ".")))
    except (ValueError, TypeError):
        return False
