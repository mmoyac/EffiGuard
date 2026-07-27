import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl
import openpyxl.styles
import openpyxl.utils
from fastapi import APIRouter, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select

from app.core.dependencies import CurrentToken, DBSession
from app.core.uid import generar_uid
from app.models.asset import Asset
from app.models.asset_family import AssetFamily
from app.models.asset_state import AssetState
from app.models.ubicacion import Ubicacion
from app.repositories.asset import AssetRepository
from app.repositories.ubicacion import UbicacionRepository
from app.schemas.asset import UNIDADES_VALIDAS

router = APIRouter(prefix="/assets/import", tags=["Assets Import"])

# El importador lee las celdas por índice: las columnas nuevas SIEMPRE se agregan
# al final. Insertarlas en medio rompería en silencio los archivos ya descargados.
_COLUMNS = [
    "uid_fisico",
    "nombre",
    "familia",
    "estado",
    "stock_actual",
    "stock_minimo",
    "valor_reposicion",
    "dias_max_prestamo",
    "proxima_mantencion",
    "ubicacion_rack",
    "ubicacion_nivel",
    "ubicacion_posicion",
    "unidad",
    "codigo_fabricante",
    "contenido_por_empaque",
    "nombre_empaque",
    "precio_compra",
]

_EXAMPLES = [
    ["", "Taladro Bosch GBH 2-26", "Herramientas Eléctricas", "Disponible", "", "", "85000", "7", "", "3", "5", "11", "unidad", "7891234567890", "", "", ""],
    ["", "Tornillo Zincado M3.5x15", "EPP Consumibles",       "Disponible", "9000", "1000", "12", "", "", "1", "2", "4", "unidad", "5001216328235", "100", "caja", "120"],
    ["", "Luces LED",              "EPP Consumibles",         "Disponible", "250.5", "50", "1200", "", "", "2", "1", "7", "metro", "", "100", "rollo", "1200"],
]


@router.get("/template")
async def download_template(token: CurrentToken, session: DBSession):
    """
    Descarga el template Excel.
    - Si el tenant ya tiene activos → los exporta pre-cargados (para editar y reimportar).
    - Si no tiene activos → exporta el template vacío con filas de ejemplo.
    """
    # Cargar familias del tenant (id → nombre)
    families_result = await session.execute(
        select(AssetFamily).where(AssetFamily.tenant_id == token.tenant_id)
    )
    families_by_id: dict[int, str] = {f.id: f.nombre for f in families_result.scalars().all()}

    # Cargar estados globales (id → nombre)
    states_result = await session.execute(select(AssetState))
    states_by_id: dict[int, str] = {s.id: s.nombre for s in states_result.scalars().all()}

    # Cargar ubicaciones del tenant (id → terna)
    ubic_result = await session.execute(
        select(Ubicacion).where(Ubicacion.tenant_id == token.tenant_id)
    )
    ubic_by_id: dict[int, Ubicacion] = {u.id: u for u in ubic_result.scalars().all()}

    # Cargar activos del tenant (solo raíz — sin hijos de kits para no duplicar)
    assets_result = await session.execute(
        select(Asset)
        .where(Asset.tenant_id == token.tenant_id)
        .where(Asset.parent_asset_id == None)  # noqa: E711
        .order_by(Asset.id)
    )
    assets = list(assets_result.scalars().all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activos"

    # Estilos
    bold = openpyxl.styles.Font(bold=True)
    header_fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = openpyxl.styles.Font(bold=True, color="E5E7EB")

    for col, header in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    if assets:
        for row_idx, asset in enumerate(assets, start=2):
            familia_nombre = families_by_id.get(asset.family_id, "")
            estado_nombre = states_by_id.get(asset.estado_id, "Disponible")
            ws.cell(row=row_idx, column=1, value=asset.uid_fisico)
            ws.cell(row=row_idx, column=2, value=asset.nombre or "")
            ws.cell(row=row_idx, column=3, value=familia_nombre)
            ws.cell(row=row_idx, column=4, value=estado_nombre)
            ws.cell(row=row_idx, column=5, value=asset.stock_actual)
            ws.cell(row=row_idx, column=6, value=asset.stock_minimo)
            ws.cell(row=row_idx, column=7, value=float(asset.valor_reposicion) if asset.valor_reposicion else "")
            ws.cell(row=row_idx, column=8, value=asset.dias_max_prestamo or "")
            ws.cell(row=row_idx, column=9, value=str(asset.proxima_mantencion) if asset.proxima_mantencion else "")
            ubic = ubic_by_id.get(asset.ubicacion_id) if asset.ubicacion_id else None
            ws.cell(row=row_idx, column=10, value=ubic.rack if ubic else "")
            ws.cell(row=row_idx, column=11, value=ubic.nivel if ubic else "")
            ws.cell(row=row_idx, column=12, value=ubic.posicion if ubic else "")
            ws.cell(row=row_idx, column=13, value=asset.unidad or "unidad")
            ws.cell(row=row_idx, column=14, value=asset.codigo_fabricante or "")
            ws.cell(row=row_idx, column=15, value=float(asset.contenido_por_empaque) if asset.contenido_por_empaque else "")
            ws.cell(row=row_idx, column=16, value=asset.nombre_empaque or "")
            ws.cell(row=row_idx, column=17, value=float(asset.precio_compra) if asset.precio_compra else "")
    else:
        for row_idx, row_data in enumerate(_EXAMPLES, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    widths = [30, 25, 22, 15, 13, 13, 16, 16, 20, 15, 15, 17, 10, 20, 21, 15, 15]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=activos.xlsx"},
    )


@router.post("")
async def import_assets(
    token: CurrentToken,
    session: DBSession,
    file: UploadFile = File(...),
    dry_run: bool = Query(False, description="Si true, valida sin modificar datos"),
):
    """
    Importa/actualiza activos desde un archivo .xlsx (upsert por uid_fisico).
    - uid_fisico vacío → crea nuevo activo con UID generado automáticamente.
    - uid_fisico existente (del mismo tenant) → actualiza el activo.
    - uid_fisico existente de otro tenant → error de fila.
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo debe ser .xlsx")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se pudo leer el archivo Excel")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=False))

    if len(rows) < 2:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo no tiene filas de datos")

    # Pre-cargar familias del tenant (nombre lower → id)
    families_result = await session.execute(
        select(AssetFamily).where(AssetFamily.tenant_id == token.tenant_id)
    )
    families: dict[str, int] = {f.nombre.strip().lower(): f.id for f in families_result.scalars().all()}

    # Pre-cargar estados globales (nombre lower → id)
    states_result = await session.execute(select(AssetState))
    states: dict[str, int] = {s.nombre.strip().lower(): s.id for s in states_result.scalars().all()}

    # Activos existentes de ESTE tenant (uid → Asset) para upsert
    tenant_assets_result = await session.execute(
        select(Asset).where(Asset.tenant_id == token.tenant_id)
    )
    tenant_assets: dict[str, Asset] = {a.uid_fisico: a for a in tenant_assets_result.scalars().all()}

    # Sólo los UIDs de ESTE tenant: desde la migración 017 la unicidad es por
    # tenant, así que un código usado por otro cliente ya no es un conflicto.
    tenant_uids: set[str] = set(tenant_assets)

    to_create: list[dict] = []
    to_update: list[tuple[Asset, dict]] = []
    errores: list[dict] = []
    uids_en_archivo: set[str] = set()

    # ── Resolución de ubicaciones ────────────────────────────────────────────
    ubic_repo = UbicacionRepository(session, token.tenant_id)
    ubicaciones_cache: dict[tuple[str, str, str], int] = {}
    ubicaciones_creadas: set[tuple[str, str, str]] = set()

    async def resolver_ubicacion(terna: tuple[str, str, str]) -> int | None:
        """Resuelve la terna contra el catálogo y la crea si no existe.

        Se crea en vez de rechazar la fila porque una ubicación es una etiqueta de
        dónde está algo, no configuración con significado de negocio como la
        familia o el estado. En dry_run no se persiste nada: sólo se cuenta.
        """
        clave = tuple(v.strip().upper() for v in terna)
        if clave in ubicaciones_cache:
            return ubicaciones_cache[clave]

        existente = await ubic_repo.get_by_posicion(*clave)
        if existente:
            ubicaciones_cache[clave] = existente.id
            return existente.id

        ubicaciones_creadas.add(clave)
        if dry_run:
            return None
        nueva = await ubic_repo.create(rack=clave[0], nivel=clave[1], posicion=clave[2])
        ubicaciones_cache[clave] = nueva.id
        return nueva.id

    for row_idx, row in enumerate(rows[1:], start=2):
        def cell_value(idx: int) -> str | None:
            if idx >= len(row):
                return None
            val = row[idx].value
            if val is None:
                return None
            s = str(val).strip()
            return s if s else None

        def cell_raw(idx: int):
            if idx >= len(row):
                return None
            return row[idx].value

        uid_fisico = cell_value(0)
        nombre = cell_value(1)
        familia_nombre = cell_value(2)
        estado_nombre = cell_value(3) or "Disponible"
        stock_actual_raw = cell_value(4)
        stock_minimo_raw = cell_value(5)
        valor_reposicion_raw = cell_value(6)
        dias_max_raw = cell_value(7)
        proxima_raw = cell_raw(8)
        # Columnas nuevas: los archivos del template anterior no las traen y
        # cell_value() devuelve None para índices fuera de rango.
        rack_raw = cell_value(9)
        nivel_raw = cell_value(10)
        posicion_raw = cell_value(11)
        unidad_raw = cell_value(12)
        codigo_fab_raw = cell_value(13)
        contenido_raw = cell_value(14)
        nombre_empaque_raw = cell_value(15)
        precio_raw = cell_value(16)

        # Ignorar filas completamente vacías
        if uid_fisico is None and nombre is None and familia_nombre is None:
            continue

        if not nombre:
            errores.append({"fila": row_idx, "motivo": "nombre es obligatorio"})
            continue

        if not familia_nombre:
            errores.append({"fila": row_idx, "motivo": "familia es obligatoria"})
            continue

        family_id = families.get(familia_nombre.strip().lower())
        if family_id is None:
            errores.append({"fila": row_idx, "motivo": f"familia '{familia_nombre}' no existe en el sistema"})
            continue

        state_id = states.get(estado_nombre.strip().lower())
        if state_id is None:
            opciones = ", ".join(sorted(states.keys()))
            errores.append({"fila": row_idx, "motivo": f"estado '{estado_nombre}' inválido. Opciones: {opciones}"})
            continue

        # Parsear campos numéricos. Todos usan False como sentinela de error:
        # None es un default legítimo (dias_max_prestamo vacío = sin límite).
        stock_actual = _parse_cantidad(stock_actual_raw)
        if stock_actual is False:
            errores.append({"fila": row_idx, "motivo": f"stock_actual '{stock_actual_raw}' no es un número válido"})
            continue

        stock_minimo = _parse_cantidad(stock_minimo_raw)
        if stock_minimo is False:
            errores.append({"fila": row_idx, "motivo": f"stock_minimo '{stock_minimo_raw}' no es un número válido"})
            continue

        valor_reposicion = _parse_decimal(valor_reposicion_raw)
        if valor_reposicion is False:
            errores.append({"fila": row_idx, "motivo": f"valor_reposicion '{valor_reposicion_raw}' no es un número válido"})
            continue

        dias_max_prestamo = _parse_int(dias_max_raw, None)
        if dias_max_prestamo is False:
            errores.append({"fila": row_idx, "motivo": f"dias_max_prestamo '{dias_max_raw}' no es un entero válido"})
            continue

        proxima_mantencion = _parse_date(proxima_raw)
        if proxima_mantencion is False:
            errores.append({"fila": row_idx, "motivo": f"proxima_mantencion '{proxima_raw}' debe tener formato YYYY-MM-DD"})
            continue

        unidad = (unidad_raw or "unidad").strip().lower()
        if unidad not in UNIDADES_VALIDAS:
            errores.append({"fila": row_idx, "motivo": f"unidad '{unidad_raw}' inválida. Opciones: {', '.join(UNIDADES_VALIDAS)}"})
            continue

        precio_compra = _parse_cantidad(precio_raw, None)
        if precio_compra is False:
            errores.append({"fila": row_idx, "motivo": f"precio_compra '{precio_raw}' no es un número válido"})
            continue
        if precio_compra is not None and precio_compra <= 0:
            errores.append({"fila": row_idx, "motivo": "precio_compra debe ser mayor a 0"})
            continue

        contenido_por_empaque = _parse_cantidad(contenido_raw, None)
        if contenido_por_empaque is False:
            errores.append({"fila": row_idx, "motivo": f"contenido_por_empaque '{contenido_raw}' no es un número válido"})
            continue
        if contenido_por_empaque is not None and contenido_por_empaque <= 0:
            errores.append({"fila": row_idx, "motivo": "contenido_por_empaque debe ser mayor a 0"})
            continue

        # Ubicación: la terna va completa o no va. Media ubicación no ubica nada.
        terna = (rack_raw, nivel_raw, posicion_raw)
        if any(terna) and not all(terna):
            errores.append({"fila": row_idx, "motivo": "la ubicación requiere rack, nivel y posición"})
            continue

        ubicacion_id = await resolver_ubicacion(terna) if all(terna) else None

        data = {
            "nombre": nombre,
            "family_id": family_id,
            "estado_id": state_id,
            "ubicacion_id": ubicacion_id,
            "unidad": unidad,
            # No es único: las unidades del mismo producto lo comparten
            "codigo_fabricante": codigo_fab_raw.strip().upper() if codigo_fab_raw else None,
            "contenido_por_empaque": contenido_por_empaque,
            "nombre_empaque": nombre_empaque_raw.strip().lower() if nombre_empaque_raw else None,
            "precio_compra": precio_compra,
            "stock_actual": stock_actual,
            "stock_minimo": stock_minimo,
            "valor_reposicion": valor_reposicion,
            "dias_max_prestamo": dias_max_prestamo,
            "proxima_mantencion": proxima_mantencion,
        }

        if not uid_fisico:
            # Nuevo activo — generar UID
            uid_fisico = generar_uid(tenant_uids | uids_en_archivo)
            uids_en_archivo.add(uid_fisico)
            to_create.append({"uid_fisico": uid_fisico, **data})

        elif uid_fisico in tenant_assets:
            # Activo existente de este tenant → actualizar
            if uid_fisico in uids_en_archivo:
                errores.append({"fila": row_idx, "motivo": f"uid_fisico '{uid_fisico}' está duplicado en el archivo"})
                continue
            uids_en_archivo.add(uid_fisico)
            to_update.append((tenant_assets[uid_fisico], data))

        else:
            # UID nuevo no existente en el sistema → crear con ese UID
            if uid_fisico in uids_en_archivo:
                errores.append({"fila": row_idx, "motivo": f"uid_fisico '{uid_fisico}' está duplicado en el archivo"})
                continue
            uids_en_archivo.add(uid_fisico)
            to_create.append({"uid_fisico": uid_fisico, **data})

    creados = 0
    actualizados = 0

    if not dry_run:
        repo = AssetRepository(session, token.tenant_id)
        for asset_data in to_create:
            await repo.create(**asset_data)
        creados = len(to_create)

        for asset_obj, update_data in to_update:
            await repo.update(asset_obj, **update_data)
        actualizados = len(to_update)

        await session.commit()

    return {
        "dry_run": dry_run,
        "creados": creados if not dry_run else 0,
        "actualizados": actualizados if not dry_run else 0,
        "validados_crear": len(to_create),
        "validados_actualizar": len(to_update),
        # Se informa para que un error de tipeo masivo en el archivo sea visible:
        # sin este dato, el catálogo crecería en silencio.
        "ubicaciones_creadas": len(ubicaciones_creadas),
        "ubicaciones_creadas_detalle": sorted(" · ".join(k) for k in ubicaciones_creadas),
        "errores": errores,
    }


# ── helpers ──────────────────────────────────────────────────────────────────

def _parse_int(value: str | None, default):
    """False si el valor no es un entero válido.

    Antes devolvía None en el caso de error, que colisionaba con el default None
    de dias_max_prestamo: un valor inválido se guardaba como 'sin límite' en vez
    de reportarse como error de fila.
    """
    if value is None:
        return default
    try:
        return int(float(str(value).strip().replace(",", ".")))
    except (ValueError, TypeError):
        return False


def _parse_cantidad(value: str | None, default=Decimal(0)):
    """Cantidad de stock: hasta 3 decimales, acepta coma o punto. False si es inválida."""
    if value is None:
        return default
    try:
        return Decimal(str(value).strip().replace(",", ".")).quantize(Decimal("0.001"))
    except (InvalidOperation, ValueError, TypeError):
        return False


def _parse_decimal(value: str | None):
    if value is None:
        return None
    try:
        return Decimal(value.replace(",", "."))
    except (InvalidOperation, AttributeError):
        return False


def _parse_date(value) -> date | None | bool:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        return False
